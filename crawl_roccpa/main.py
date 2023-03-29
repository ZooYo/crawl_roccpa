import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook

PAGE_SIZE = 211  # 總頁數
KEYWORD = "聯合"


def get_crawl_data(keyword=KEYWORD, page_size=PAGE_SIZE):

    members = []
    for i in range(1, page_size+1):
        members.extend(get_page_data(keyword, i))

    return members


def get_page_data(keyword, page_number):
    members = []
    page_url = f"https://www.roccpa.org.tw/member_search/list2"
    page_params = {
        "AntiToken": "ywXGmdmH5F%2bpXAENs%2fsVkk4k7htDZvu585GLDotEd68%3d",
        "location": "",
        "fields": 2,
        "p": page_number,
        "keys": keyword
    }
    headers = {
        "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36 Edg/110.0.1587.69"
    }

    page_response = requests.get(page_url, params=page_params, headers=headers)
    page_soup = BeautifulSoup(page_response.content, "html.parser")

    rows = page_soup.findAll("tr")
    for row_index, row in enumerate(rows):
        if row_index != 0:
            member = {}

            cols = row.findAll("td")
            for col_index, col in enumerate(cols):
                if col_index == 1:
                    member["accountant"] = col.text
                elif col_index == 2:
                    member["office_name"] = col.text

            more_url = "https://www.roccpa.org.tw/member_search/more2"
            more_info = {}
            if row.find("a"):
                more_id = row.find("a")['href'].replace("more2?id=", "")
                more_params = {"id": more_id}
                more_response = requests.get(more_url, params=more_params, headers=headers)
                more_soup = BeautifulSoup(more_response.content, "html.parser")
                target_ol = more_soup.findAll('ol')[1]

                for more_index, li in enumerate(target_ol):
                    if li.find('b') != -1:
                        more_info[li.find('b').text] = li.find('div').text.strip()
            member["more_info"] = more_info
            members.append(member)
    return members


def write_data_to_excel_from_list(data_list):

    # Create a new workbook
    wb = Workbook()

    # Select the active worksheet
    ws = wb.active

    # Write the header row
    header = ['序號', '會計師姓名', '事務所名稱', '事務所地址', '事務所電話', '事務所傳真', 'E-mail', '會籍編號']
    for col_num, col_title in enumerate(header, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = col_title

    # Write the data rows
    for row_num, row_data in enumerate(data_list, 2):
        cell = ws.cell(row=row_num, column=1)
        cell.value = row_num-1
        cell = ws.cell(row=row_num, column=2)
        cell.value = row_data['accountant']
        cell = ws.cell(row=row_num, column=3)
        cell.value = row_data['office_name']

        more_info = row_data['more_info']
        if '事務所地址' in more_info:
            cell = ws.cell(row=row_num, column=4)
            cell.value = more_info['事務所地址']
        if '事務所電話' in more_info:
            cell = ws.cell(row=row_num, column=5)
            cell.value = more_info['事務所電話']
        if '事務所傳真' in more_info:
            cell = ws.cell(row=row_num, column=6)
            cell.value = more_info['事務所傳真']
        if 'E-mail' in more_info:
            cell = ws.cell(row=row_num, column=7)
            cell.value = more_info['E-mail']
        cell = ws.cell(row=row_num, column=8)
        cell.value = get_city_member_number(row_data['more_info'])

    # Save the workbook
    wb.save('output.xlsx')


def get_city_member_number(more_dict):
    for key in more_dict.keys():
        if "會籍編號" in key:
            return more_dict[key]

    return ""


if __name__ == '__main__':

    member_data = get_crawl_data()
    write_data_to_excel_from_list(member_data)

